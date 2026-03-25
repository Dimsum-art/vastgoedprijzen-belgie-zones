"""Process Statbel municipality-level real estate data into ground price-per-m2 JSON."""

import json
from pathlib import Path

import requests
from openpyxl import load_workbook

STATBEL_MUNI_URL = (
    "https://statbel.fgov.be/sites/default/files/files/opendata/"
    "Verkoop%20van%20onroerende%20goederen%20per%20gemeente%20%282010-2019%29/"
    "immo_by_municipality_2010-2019.xlsx"
)

# 2019 Belgian municipality mergers and Hainaut NIS code renumbering.
# Maps old (pre-2019) NIS codes to current codes.
NIS_MERGE_MAP: dict[str, str] = {
    # Flemish mergers
    "12030": "12041",  # Puurs -> Puurs-Sint-Amands
    "12034": "12041",  # Sint-Amands -> Puurs-Sint-Amands
    "44001": "44084",  # Aalter (old) -> Aalter (new)
    "44029": "44084",  # Knesselare -> Aalter (new)
    "44011": "44083",  # Deinze (old) -> Deinze (new)
    "44049": "44083",  # Nevele -> Deinze (new)
    "44036": "44085",  # Lovendegem -> Lievegem
    "44072": "44085",  # Waarschoot -> Lievegem
    "44080": "44085",  # Zomergem -> Lievegem
    "45017": "45068",  # Kruishoutem -> Kruisem
    "45057": "45068",  # Zingem -> Kruisem
    "71047": "72042",  # Opglabbeek -> Oudsbergen
    "72025": "72042",  # Meeuwen-Gruitrode -> Oudsbergen
    "72029": "72043",  # Overpelt -> Pelt
    "72040": "72043",  # Neerpelt -> Pelt
    # Hainaut arrondissement renumbering
    "52043": "51067",  # Enghien -> Edingen
    "52063": "51068",  # Silly -> Opzullik
    "54007": "57096",  # Mouscron -> Moeskroen
    "54010": "57097",  # Comines-Warneton -> Komen-Waasten
    "55010": "55085",  # Seneffe
    "55022": "58001",  # La Louviere
    "55023": "51069",  # Lessines -> Lessen
    "55039": "55086",  # Manage
    "56011": "58002",  # Binche
    "56085": "58003",  # Estinnes
    "56087": "58004",  # Morlanwelz
}

DATA_DIR = Path(__file__).parent.parent / "data"
XLSX_PATH = DATA_DIR / "immo_by_municipality_2010-2019.xlsx"
OUTPUT_PATH = DATA_DIR / "prices_grond_muni.json"


def download_muni_data() -> None:
    """Download Statbel municipality XLSX if not present."""
    if XLSX_PATH.exists():
        print(f"Municipality data already downloaded: {XLSX_PATH}")
        return
    print("Downloading Statbel municipality data...")
    resp = requests.get(STATBEL_MUNI_URL, timeout=120)
    resp.raise_for_status()
    XLSX_PATH.write_bytes(resp.content)
    print(f"Saved to {XLSX_PATH}")


def compute_muni_prices(xlsx_path: Path) -> dict[str, int]:
    """Parse municipality XLSX -> {nis5: price_per_m2}.

    Aggregates total price and total surface across all property types
    and years per municipality, then computes round(total_price / total_surface).
    """
    wb = load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    # Detect column indices by keyword matching
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h in ("cd_refnis",) or h == "refnis_munty":
            col_map["nis"] = i
        elif "total_price" in h or "totale_prijs" in h:
            col_map["price"] = i
        elif "total_surface" in h or "totale_oppervlakte" in h:
            col_map["surface"] = i

    required = {"nis", "price", "surface"}
    missing = required - set(col_map.keys())
    if missing:
        raise ValueError(
            f"Could not find columns: {missing}. Headers found: {headers}"
        )

    # Aggregate by municipality
    agg: dict[str, dict[str, float]] = {}
    for row in rows[1:]:
        raw_nis = row[col_map["nis"]]
        price = row[col_map["price"]]
        surface = row[col_map["surface"]]

        if not raw_nis or not price or not surface:
            continue

        try:
            nis = (
                str(int(raw_nis))
                if isinstance(raw_nis, (int, float))
                else str(raw_nis).strip()
            )
            nis = NIS_MERGE_MAP.get(nis, nis)
            price_val = float(price)
            surface_val = float(surface)
        except (ValueError, TypeError):
            continue

        if surface_val <= 0:
            continue

        if nis not in agg:
            agg[nis] = {"total_price": 0.0, "total_surface": 0.0}
        agg[nis]["total_price"] += price_val
        agg[nis]["total_surface"] += surface_val

    # Compute price/m2
    result: dict[str, int] = {}
    for nis, data in agg.items():
        if data["total_surface"] > 0:
            pm2 = round(data["total_price"] / data["total_surface"])
            if pm2 > 0:
                result[nis] = pm2

    wb.close()
    return result


if __name__ == "__main__":
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    download_muni_data()

    print("Processing municipality data...")
    prices = compute_muni_prices(XLSX_PATH)
    print(f"Computed price/m2 for {len(prices)} municipalities")

    # Coverage stats
    vals = sorted(prices.values())
    print(f"Range: {vals[0]} - {vals[-1]} EUR/m2")
    print(f"Median: {vals[len(vals) // 2]} EUR/m2")
    print(f"Mean: {round(sum(vals) / len(vals))} EUR/m2")

    with open(OUTPUT_PATH, "w") as f:
        json.dump(prices, f, sort_keys=True)
    print(f"Written to {OUTPUT_PATH}")
