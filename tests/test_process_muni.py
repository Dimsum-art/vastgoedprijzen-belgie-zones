"""Tests for municipality price computation logic."""

import json
from pathlib import Path

import pytest

from scripts.process_muni_prices import NIS_MERGE_MAP, compute_muni_prices

DATA_DIR = Path(__file__).parent.parent / "data"


# --- NIS merge map tests ---

def test_nis_merge_map_values_are_strings():
    """All values in NIS_MERGE_MAP should be 5-digit string NIS codes."""
    for old, new in NIS_MERGE_MAP.items():
        assert isinstance(old, str) and len(old) == 5, f"Key {old!r} not 5-digit"
        assert isinstance(new, str) and len(new) == 5, f"Value {new!r} not 5-digit"


def test_nis_merge_map_no_self_references():
    """No NIS code should map to itself."""
    for old, new in NIS_MERGE_MAP.items():
        assert old != new, f"Self-reference: {old} -> {new}"


def test_nis_merge_map_no_chains():
    """Values should not appear as keys (no transitive merges)."""
    values = set(NIS_MERGE_MAP.values())
    keys = set(NIS_MERGE_MAP.keys())
    chains = values & keys
    assert len(chains) == 0, f"Chain merges found: {chains}"


# --- compute_muni_prices tests ---

def test_compute_muni_prices_requires_xlsx(tmp_path):
    """Should raise if XLSX is missing required columns."""
    # Create a minimal invalid XLSX
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["wrong_col1", "wrong_col2"])
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(str(xlsx_path))

    with pytest.raises(ValueError, match="Could not find columns"):
        compute_muni_prices(xlsx_path)


def test_compute_muni_prices_basic(tmp_path):
    """Basic computation with valid data."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["cd_refnis", "cd_year", "total_price", "total_surface"])
    ws.append([11001, 2015, 500000.0, 200.0])
    ws.append([11001, 2016, 600000.0, 250.0])
    ws.append([21001, 2015, 1000000.0, 100.0])
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(str(xlsx_path))

    result = compute_muni_prices(xlsx_path)
    # 11001: (500000+600000)/(200+250) = 1100000/450 = 2444
    assert result["11001"] == round(1100000 / 450)
    # 21001: 1000000/100 = 10000
    assert result["21001"] == 10000


def test_compute_muni_prices_applies_merge_map(tmp_path):
    """Old NIS codes should be merged to new codes."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["cd_refnis", "cd_year", "total_price", "total_surface"])
    # 12030 (Puurs) + 12034 (Sint-Amands) -> 12041 (Puurs-Sint-Amands)
    ws.append([12030, 2015, 300000.0, 100.0])
    ws.append([12034, 2016, 500000.0, 200.0])
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(str(xlsx_path))

    result = compute_muni_prices(xlsx_path)
    assert "12030" not in result
    assert "12034" not in result
    assert "12041" in result
    assert result["12041"] == round(800000 / 300)


def test_compute_muni_prices_skips_zero_surface(tmp_path):
    """Rows with zero surface should be skipped."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["cd_refnis", "cd_year", "total_price", "total_surface"])
    ws.append([11001, 2015, 500000.0, 0.0])
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(str(xlsx_path))

    result = compute_muni_prices(xlsx_path)
    assert len(result) == 0


# --- Integration test: real data file ---

@pytest.mark.skipif(
    not (DATA_DIR / "prices_grond_muni.json").exists(),
    reason="prices_grond_muni.json not generated yet",
)
def test_grond_muni_json_reasonable_values():
    """Municipality prices should be in a reasonable range for ground prices."""
    with open(DATA_DIR / "prices_grond_muni.json") as f:
        prices = json.load(f)
    assert len(prices) > 500, f"Expected 500+ municipalities, got {len(prices)}"
    vals = list(prices.values())
    # Ground prices typically 10-2000 EUR/m2
    assert min(vals) > 0
    assert max(vals) < 10000, f"Max {max(vals)} seems too high for ground price/m2"
